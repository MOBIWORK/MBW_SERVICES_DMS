import frappe
from frappe import _
from io import BytesIO
from mbw_dms.api.common import ( get_employee_info,gen_response)
from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from abc import ABC, abstractmethod

list_name_col_x =['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC']
week_days = {
    "Mon": "Thứ Hai",
    "Tue": "Thứ Ba",
    "Wed": "Thứ Tư",
    "Thu": "Thứ Năm",
    "Fri": "Thứ Sáu",
    "Sat": "Thứ Bảy",
    "Sun": "Chủ Nhật"
}
class MakeExcelFile(ABC):
    # Định dạng chung
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_tableFill = PatternFill(start_color="46ea46", end_color="46ea46", fill_type="solid")
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center")
    # Font and alignment settings
    title_font = Font(bold=True, size=14)
    subtitle_font = Font(italic=True, size=10)
    header_font = Font(size=11, bold=False, color="5C6B82")  # Define a custom font color
    bold_font = Font(size=11, bold=True, color="000000")  # Bold font for subheaders
    center_alignment = Alignment(horizontal="center", vertical="center")  # Center alignment
    left_alignment = Alignment(horizontal="left", vertical="center")  # Left alignment
    border_style = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    ) 
    def __init__(self,excel_name,header_row1,header_row2,field_row,merge_header,data=[],footer=[]):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.data = data

        if bool(excel_name):
            self.excel_name = excel_name

        if bool(header_row1):
            self.header_row1 = header_row1

        if bool(header_row2):
            self.header_row2 = header_row2

        if bool(field_row):
            self.field_row = field_row

        if bool(merge_header):
            self.merge_header = merge_header


        if bool(footer):
            self.footer = footer
        
        self.convert_data()

    def addStartRowHeader(self,start_header_row):
        self.startHeader_row =start_header_row
    
    def addStartRowContent(self,start_content_row):
        self.startContent_row =start_content_row

    def addFooterData(self,data):
        self.data_footer = data

    @abstractmethod
    def make(self):
        pass

    @abstractmethod
    def convert_data(self):
        pass

    # header
    def create_header(self):
        headers_row1 = self.header_row1
        headers_row2 = self.header_row2 or []
        col_span = self.merge_header or []
        for mg in col_span:
            try:
                self.ws.unmerge_cells(mg)
            except :
                print("not merge ",mg)
        # điền title header
        # Điền các giá trị tiêu đề cho hàng đầu tiên bằng vòng lặp
        for col_idx, header in enumerate(headers_row1, start=1):
            cell = self.ws.cell(row=9, column=col_idx, value=header)
            #áp dụng style
            cell.font = self.bold_font
            cell.alignment = self.center_alignment
            cell.border = self.border_style
            cell.fill = self.header_tableFill

        
        # Fill in the header values for the second row using a loop
        if len(headers_row2) > 0:
            for col_idx, header in enumerate(headers_row2, start=1):
                cell = self.ws.cell(row=11, column=col_idx, value=header)
                #áp dụng style
                cell.font = self.bold_font  # Use bold font for the subheader
                cell.alignment = self.center_alignment
                cell.border = self.border_style
                cell.fill = self.header_tableFill
        # # gom các cột cha với các cột con
        if len(col_span):
            for mg in col_span:
                self.ws.merge_cells(mg)
        print("make header xong")
        return self

    # content
    def create_content(self):
        sort_list = self.field_row
        # print("sort_list",sort_list)
        # Add the data to the worksheet
        start_row = self.startContent_row
        # ghi nội dung vào bản
        for row_idx, row_data in enumerate(self.data, start=start_row):
            cell = self.ws.cell(row=row_idx, column=1, value=row_idx-11)
            cell.alignment = self.center_alignment
            cell.border = self.border_style
            for col_idx, value in enumerate(sort_list, start=2):
                cell = self.ws.cell(row=row_idx, column=col_idx, value=row_data.get(value) or 0)
                cell.alignment = self.left_alignment if col_idx <= 4 else self.center_alignment
                cell.border = self.border_style

    # footer
    def create_footer(self):
        start_footer = self.startContent_row  + len(self.data )
        if self.data_footer:
            sort_footer = self.footer
            if sort_footer:
                # for row_idx, row_data in enumerate(self.data_footer, start=start_row):
                
                for col_idx, value in enumerate(sort_footer, start=1):
                    # if col_idx == 1:
                    #     cell = self.ws.cell(row=row_idx, column=col_idx, value=row_idx)
                    #     cell.alignment = self.center_alignment
                    # else:
                    cell = self.ws.cell(row=start_footer, column=col_idx, value=self.data_footer.get(value))
                    cell.alignment = self.left_alignment if col_idx <= 4 else self.center_alignment
                    cell.border = self.border_style
                cell = self.ws.cell(row=start_footer, column=2, value="Tổng")
                cell.alignment = self.center_alignment
                cell.border = self.border_style
        return self