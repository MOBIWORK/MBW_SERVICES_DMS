import frappe
from frappe import _
from io import BytesIO
from mbw_dms.api.common import ( get_employee_info,gen_response)
from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


name_report = {
    "Report KPI":"BÁO CÁO KPI",
    "Report Checkin": "BÁO CÁO VIẾNG THĂM KHÁCH HÀNG",
    "Report Inventory": "BÁO CÁO TỒN SẢN PHẨM THEO KHÁCH HÀNG",
    "Report Sell": "BÁO CÁO TỔNG HỢP BÁN HÀNG",
    "Report Order": "BÁO CÁO TỔNG HỢP ĐẶT HÀNG",
    "Report Customer": "BÁO CÁO KHÁCH HÀNG MỚI",
    "Report Customer Checkin": "BÁO CÁO THỐNG KÊ KHÁCH HÀNG VIẾNG THĂM LẦN ĐẦU"
}
list_name_col_x =['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC']
week_days = {
    "Mon": "Thứ Hai",
    "Tue": "Thứ Ba",
    "Wed": "Thứ Tư",
    "Thu": "Thứ Năm",
    "Fri": "Thứ Sáu",
    "Sat": "Thứ Bảy",
    "Sun": "Chủ Nhật"
}
columnReport = {
    "Report KPI" : {
        "column" : {
            "main_columns": ["STT", "Mã nhân viên","Nhân viên", "Nhóm bán hàng","Số khách hàng viếng thăm","","", "Số khách hàng viếng thăm duy nhất","","", "Số khách đặt hàng","","","Số khách hàng thêm mới","","","Số đơn hàng","","", "Doang số (VNĐ)","","","Doang thu (VNĐ)","","", "Sản Lượng","","","SKU","","", "Số giờ làm việc""","",],
            "sub_column": ["","","","","KH","TH","TL(%)","KH","TH","TL(%)","KH","TH","TL(%)","KH","TH","TL(%)","KH","TH","TL(%)","KH","TH","TL(%)","KH","TH","TL(%)","KH","TH","TL(%)","KH","TH","TL(%)","KH","TH","TL(%)",],
        },
        "col_span": ["A9:A11", "B9:B11", "C9:C11", "D9:D11","E9:G10","H9:J10","K9:M10","N9:P10","Q9:S10","T9:V10","W9:Y10","Z9:AB10","AC9:AE10","AF9:AH10"],
        # "col_span": ["A9:A10", "B9:B10", "C9:C10", "D9:D10","E9:G9","H9:J9","K9:M9","N9:P9","Q9:S9","T9:V9","W9:Y9","Z9:AB9","AC9:AE9","AF9:AH9"],
        "show_data": ["nhan_vien_ban_hang","ten_nv","nhom_ban_hang","kh_vt","th_vt","tl_vt","kh_vt_dn","th_vt_dn","tl_vt_dn","kh_dat_hang","th_dat_hang","tl_dat_hang","kh_kh_moi","th_kh_moi","tl_kh_moi","kh_don_hang","th_don_hang","tl_don_hang","kh_doanh_so","th_doanh_so","tl_don_hang","kh_doanh_thu","th_doanh_thu","tl_doanh_thu","kh_san_lg","th_san_lg","tl_san_luong","kh_sku","th_sku","tl_sku","kh_so_gio_lam_viec","th_so_gio_lam_viec","tl_so_gio_lam_viec"],
        "content_start_at": 12,
        "footer" : ["","","","","tong_kh_vt","tong_th_vt","","tong_kh_vt_dn","tong_th_vt_dn","","tong_kh_dat_hang","tong_th_dat_hang","","tong_kh_kh_moi","tong_th_kh_moi","","tong_kh_don_hang","tong_th_don_hang","","tong_kh_doanh_so","tong_th_doanh_so","","tong_kh_doanh_thu","tong_th_doanh_thu","","tong_kh_san_lg","tong_th_san_lg","","tong_kh_sku","tong_th_sku","","tong_kh_so_gio_lam_viec","tong_th_so_gio_lam_viec",""],
        "column_widths" : {
                            "A": 5, "B": 15, "C": 25, "D": 35,
                            "E": 10, "F": 10, "G": 10, "H": 10,
                            "I": 10, "J": 10, "K": 10, "L": 10,
                            "M": 10, "N": 10, "O": 10, "P": 10,
                            "Q": 10, "R": 10, "S": 10, "T": 10,
                            "U": 10, "V": 10, "W": 10, "X": 10,
                            "Y": 10, "Z": 10, "AA": 10, "AB": 10,
                            "AC": 10, "AD": 10, "AE": 10, "AF": 10,
                            "AG": 10, "AH": 10
                        }
    },
    "Report Checkin" : {
        "column" : {
            "main_columns": ["STT", "Mã nhân viên","Nhân viên", "Nhóm bán hàng","Ngày","Thứ","Giờ làm", "Giờ VT","Khách hàng","","","","","","","Viếng Thăm","","","","","","","","","","","Số km tự động (km)","Số km tự di chuyển (km)","Vận tốc (km/h)"],
            "sub_column": ["","","","","","","","","Mã KH","Tên KH","Liên hệ","Loại KH","Nhóm KH","SĐT","Địa chỉ","Checkin","Checkout","Số giờ VT","Địa chỉ checkin","Khoảng cách (km)","Thiết bị","Số ảnh chụp","Đúng tuyến","Đơn hàng","Ghi tồn","Ghi chú","","",""],
        },
        "col_span": ["A9:A11", "B9:B11", "C9:C11", "D9:D11","E9:E11","F9:F11","G9:G11","H9:H11","I9:O10","P9:Z10","AA9:AA11","AB9:AB11","AC9:AC11"],
        # "col_span": [],
        "show_data":[ "employee_code","employee_name", "sale_group","time","day_week","total_work", "total_time","customer_code","customer_name","customer_contact","customer_type","customer_group","customer_sdt","customer_address","checkin","checkout","time_check","checkin_address","distance","device","total_image","is_router","is_order","is_check_inventory","ghi chú","","",""],
        "content_start_at": 12,
        "column_widths" : {
                            "A": 5, "B": 15, "C": 25, "D": 35,
                            "E": 10, "F": 10, "G": 10, "H": 10,
                            "I": 10, "J": 10, "K": 10, "L": 10,
                            "M": 10, "N": 10, "O": 10, "P": 10,
                            "Q": 10, "R": 10, "S": 10, "T": 10,
                            "U": 10, "V": 10, "W": 10, "X": 10,
                            "Y": 10, "Z": 10, "AA": 10, "AB": 10,
                            "AC": 10, "AD": 10, "AE": 10, "AF": 10,
                            "AG": 10, "AH": 10
                        }
    },
    "Report Inventory" : {
        "column" : {
            "main_columns": ["STT","Mã khách hàng","Tên khách hàng","Loại khách hàng","Địa chỉ","Mã sản phẩm","Tên sản phẩm","Hạn sử dụng","ĐVT","Tồn","Giá sản phẩm","Tổng giá trị","Ngày cập nhật","Mã nhân viên","Nhân viên cập nhật"],
        },
        "show_data":["customer_code","customer_name","customer_type","customer_address","item_code","item_name","exp_time","item_unit","quantity","item_price","total_quantity","update_at","update_bycode","update_byname"],
        "content_start_at": 10,
        "column_widths" : {
                            "A": 5, "B": 15, "C": 25, "D": 35,
                            "E": 10, "F": 10, "G": 10, "H": 10,
                            "I": 10, "J": 10, "K": 10, "L": 10,
                            "M": 10, "N": 10, "O": 10, "P": 10,
                            "Q": 10, "R": 10, "S": 10, "T": 10,
                            "U": 10, "V": 10, "W": 10, "X": 10,
                            "Y": 10, "Z": 10, "AA": 10, "AB": 10,
                            "AC": 10, "AD": 10, "AE": 10, "AF": 10,
                            "AG": 10, "AH": 10
                        }
    },
    "Report Sell":{
        "column": {
            "main_columns": ['STT', 'Đơn bán', 'Khách hàng', 'Khu vực', 'Ngày tạo', 'Nhân Viên', 'Mã sản phẩm', 'Tên sản phẩm', 'Mã kho', 'Tên kho', 'Nhóm sản phẩm', 'Nhãn hàng', 'Số lượng', 'ĐVT', 'Đơn giá', 'Chiết khấu(%)', 'Tiền chiết khấu(VNĐ)', 'Tổng tiền(VNĐ)', 'Thành tiền(VNĐ)', 'Tiền VAT(VNĐ)', 'Chiết khấu(VNĐ)', 'Tổng tiền(VNĐ)']
        },
        "show_data": ['name', 'customer', 'territory', 'posting_date', 'sales_person', 'item_code', 'item_name', 'set_warehouse', 'name_warehouse', 'item_group', 'brand', 'qty', 'item_unit', 'rate', 'discount_percentage', 'discount_amount', 'amount', 'total', 'tax_amount', 'discount_amount', 'grand_total'],
        "content_start_at": 10,
        "footer" : ["","","","","","","","","","","","","sum_qty","","","","sum_discount_items","sum_grand_items","sum_total","sum_vat","sum_discount_amount","sum_grand_total"],
        "column_widths" : {
                            "A": 5, "B": 15, "C": 25, "D": 35,
                            "E": 10, "F": 10, "G": 10, "H": 10,
                            "I": 10, "J": 10, "K": 10, "L": 10,
                            "M": 10, "N": 10, "O": 10, "P": 10,
                            "Q": 10, "R": 10, "S": 10, "T": 10,
                            "U": 10, "V": 10, "W": 10, "X": 10,
                            "Y": 10, "Z": 10, "AA": 10, "AB": 10,
                            "AC": 10, "AD": 10, "AE": 10, "AF": 10,
                            "AG": 10, "AH": 10
                        }
    },
    "Report Order":{
       "column": {
            "main_columns": ['STT', 'Đơn bán', 'Khách hàng', 'Khu vực', 'Ngày tạo', 'Nhân Viên', 'Mã sản phẩm', 'Tên sản phẩm', 'Mã kho', 'Tên kho', 'Nhóm sản phẩm', 'Nhãn hàng', 'Số lượng', 'ĐVT', 'Đơn giá', 'Chiết khấu(%)', 'Tiền chiết khấu(VNĐ)', 'Tổng tiền(VNĐ)', 'Thành tiền(VNĐ)', 'Tiền VAT(VNĐ)', 'Chiết khấu(VNĐ)', 'Tổng tiền(VNĐ)']
        },
        "show_data": ['name', 'customer', 'territory', 'transaction_date', 'sales_person', 'item_code', 'item_name', 'set_warehouse', 'name_warehouse', 'item_group', 'brand', 'qty', 'item_unit', 'rate', 'discount_percentage', 'discount_amount', 'amount', 'total', 'tax_amount', 'discount_amount', 'grand_total'],
        "content_start_at": 10,
        "footer" : ["","","","","","","","","","","","","sum_qty","","","","sum_discount_items","sum_grand_items","sum_total","sum_vat","sum_discount_amount","sum_grand_total"],
        "column_widths" : {
                            "A": 5, "B": 15, "C": 25, "D": 35,
                            "E": 10, "F": 10, "G": 10, "H": 10,
                            "I": 10, "J": 10, "K": 10, "L": 10,
                            "M": 10, "N": 10, "O": 10, "P": 10,
                            "Q": 10, "R": 10, "S": 10, "T": 10,
                            "U": 10, "V": 10, "W": 10, "X": 10,
                            "Y": 10, "Z": 10, "AA": 10, "AB": 10,
                            "AC": 10, "AD": 10, "AE": 10, "AF": 10,
                            "AG": 10, "AH": 10
                        }
    },
    "Report Customer": {},
    "Report Customer Checkin": {}


}
class MakeExcel :
    # Định dạng chung
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_tableFill = PatternFill(start_color="46ea46", end_color="46ea46", fill_type="solid")
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center")
    # Font and alignment settings
    title_font = Font(bold=True, size=14)
    subtitle_font = Font(italic=True, size=10)
    header_font = Font(size=11, bold=False, color="5C6B82")  # Define a custom font color
    bold_font = Font(size=11, bold=True, color="000000")  # Bold font for subheaders
    center_alignment = Alignment(horizontal="center", vertical="center")  # Center alignment
    left_alignment = Alignment(horizontal="left", vertical="center")  # Left alignment
    border_style = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    ) 

    def __init__(self,report_type,data):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.report_type = report_type
        self.data_content = data.get("data")
        self.data_footer = data.get("sum") or False
        self.description_data = []
        self.org_info(report_type)

    # lấy thông tin tổ chức 
    def org_info(self,report_type):
        if len(self.description_data) == 0 :
            try:
                employee_info = get_employee_info()
                if employee_info:
                    employee_info = frappe._dict(employee_info)
                    if not employee_info.company :
                        return gen_response(500,_("Account not in any company!"))
                    company_info = frappe.db.get_value("Company",employee_info.company,["name","phone_no"],as_dict=1)
                    address = frappe.db.get_all(doctype= "Address",filters= {"link_doctype": "Company", "link_name" : employee_info.company},fields=["*"])
                    address_title = address[0].address_title if len(address) >0 else ""
                    phone = company_info.phone_no if company_info.phone_no != None else ""
                    show_time = self.line_time() if self.line_time() else []
                    self.description_data = [
                        ["", employee_info.company , "", "", "", ""],
                        ["", f"Điện thoại: {phone}", "", "", "", ""],
                        ["", f"Địa chỉ: {address_title}", "", "", "", ""],
                        ["", "", "", "", "", ""],
                        ["", name_report[report_type], "", "", "", ""],
                        show_time
                    ]

                    print("self.description_data",self.description_data)
            except Exception as e :
                print("Lỗi khi tạo ",e)
        return self
    # tạo hiển thị kiểu thời gian
    def line_time(self):
        return []
    #tạo bảng report hoàn chỉnh
    def make(self):
        if not  self.make_description_header():
            self.make_description_header()
        self.make_header()
        if len(self.data_content) > 0:
            self.make_table()
        self.make_footer()
        xlsx_file = BytesIO()
        self.wb.save(xlsx_file)
        # # xử lý gửi excel - xóa trên server
        return xlsx_file

    # tạo phần tiêu đề report
    def make_description_header(self):
        merge = ['B1:F1','B2:F2','B3:F3','B5:F5',"B6:F6","B7:F7"]
        # Populate the worksheet with data
        for row in self.description_data :
            self.ws.append(row)
        # Merge cells for styling as in the image
        for mer in merge:
            self.ws.merge_cells(mer)
        self.ws['B1'].font = self.bold_font
        self.ws['B1'].alignment = self.center_alignment
        self.ws['B1'].fill = self.header_fill

        self.ws['B5'].font = self.title_font
        self.ws['B5'].alignment = self.center_alignment

        self.ws['B6'].font = self.subtitle_font
        self.ws['B6'].alignment = self.center_alignment

        self.ws['B7'].font = self.subtitle_font
        self.ws['B7'].alignment = self.center_alignment
        # Định dạng chiều rộng cột
        column_widths = columnReport[self.report_type]["column_widths"] 
        for col, width in column_widths.items():
            self.ws.column_dimensions[col].width = width
        return self
    
    # tạo header bảng
    def make_header(self):
        headers_row1 = columnReport[self.report_type]["column"]["main_columns"]
        headers_row2 = columnReport[self.report_type]["column"].get("sub_column") or []
        col_span = columnReport[self.report_type].get("col_span") or []
        for mg in col_span:
            try:
                self.ws.unmerge_cells(mg)
            except :
                print("not merge ",mg)
        # điền title header
        # Điền các giá trị tiêu đề cho hàng đầu tiên bằng vòng lặp
        for col_idx, header in enumerate(headers_row1, start=1):
            cell = self.ws.cell(row=9, column=col_idx, value=header)
            #áp dụng style
            cell.font = self.bold_font
            cell.alignment = self.center_alignment
            cell.border = self.border_style
            cell.fill = self.header_tableFill

        
        # Fill in the header values for the second row using a loop
        if len(headers_row2) > 0:
            for col_idx, header in enumerate(headers_row2, start=1):
                cell = self.ws.cell(row=11, column=col_idx, value=header)
                #áp dụng style
                cell.font = self.bold_font  # Use bold font for the subheader
                cell.alignment = self.center_alignment
                cell.border = self.border_style
                cell.fill = self.header_tableFill
        # # gom các cột cha với các cột con
        if len(col_span):
            for mg in col_span:
                self.ws.merge_cells(mg)
        print("make header xong")
        return self

    # tạo nội dung bảng
    def make_table(self):
        sort_list = columnReport[self.report_type]["show_data"]
        # print("sort_list",sort_list)
        # Add the data to the worksheet
        start_row = columnReport[self.report_type]["content_start_at"]
        # ghi nội dung vào bản
        for row_idx, row_data in enumerate(self.data_content, start=start_row):
            cell = self.ws.cell(row=row_idx, column=1, value=row_idx-11)
            cell.alignment = self.center_alignment
            cell.border = self.border_style
            for col_idx, value in enumerate(sort_list, start=2):
                cell = self.ws.cell(row=row_idx, column=col_idx, value=row_data.get(value) or 0)
                cell.alignment = self.left_alignment if col_idx <= 4 else self.center_alignment
                cell.border = self.border_style
        

        return self
    def make_footer(self):
        if self.data_footer:
            sort_footer = columnReport[self.report_type]["footer"]
            if sort_footer:
                start_row = len(self.data_content) +  (columnReport[self.report_type]["content_start_at"])
                # for row_idx, row_data in enumerate(self.data_footer, start=start_row):
                
                for col_idx, value in enumerate(sort_footer, start=1):
                    # if col_idx == 1:
                    #     cell = self.ws.cell(row=row_idx, column=col_idx, value=row_idx)
                    #     cell.alignment = self.center_alignment
                    # else:
                    cell = self.ws.cell(row=start_row, column=col_idx, value=self.data_footer.get(value))
                    cell.alignment = self.left_alignment if col_idx <= 4 else self.center_alignment
                    cell.border = self.border_style
                cell = self.ws.cell(row=start_row, column=2, value="Tổng")
                cell.alignment = self.center_alignment
                cell.border = self.border_style
        return self
    def convert_tp_to_string(self,tp):
        return  datetime.fromtimestamp(float(tp)).strftime("%d/%m/%Y")
    

class MakeExcelKpi(MakeExcel):
    def __init__(self,report_type,data,from_time,to_time):
        self.month = from_time
        self.year = to_time
        super().__init__(report_type,data)
    def line_time(self):
        return ["", f"Tháng: {self.month} - Năm: {self.year} ", "", "", "", ""]

class MakeExcelCheckin(MakeExcel):
    def __init__(self,report_type,data,from_time,to_time):
        if bool(from_time):
            self.from_time = self.convert_tp_to_string(from_time)
        if bool(to_time):
            self.to_time =  self.convert_tp_to_string(to_time)
        self.fields_group = ["employee_code","employee_name", "sale_group","time","day_week","total_work", "total_time"] 
        super().__init__(report_type,data)
        self.changer_data()
    # định nghĩa thời gian
    def line_time(self):
        if hasattr(self, 'from_time') and hasattr(self, 'to_time'):
            return ["", f"{self.from_time} - {self.to_time}", "", "", "", ""]
        elif hasattr(self, 'from_time'):
            return ["", f"Từ {self.from_time}", "", "", "", ""]
        elif hasattr(self, 'to_time'):
            return ["", f"Tính đến {self.to_time}", "", "", "", ""]
        else:
            return ["", "", "", "", "", ""]
    def changer_data(self):
        data  = self.data_content
        new_data = []
        for idx_data,checkin in enumerate(data,1):
            # xử lý thời gian ,ngày tháng
            checkin["total_work"] = round(float(checkin["total_work"])/ 60,2)
            checkin["total_time"] = round(float(checkin["total_time"])/ 60,2)
            checkin["time"] =  self.convert_tp_to_string(checkin["create_time"])
            checkin["day_week"] = week_days[datetime.fromtimestamp(float(checkin["create_time"])).strftime("%a")]

            customers = checkin.get("customers")
            for index,customer in enumerate(customers):
                customer["time_check"] = round(float(customer["time_check"])/ 60,2)
                if index == 0 :
                    data_appen = {**checkin,**customer}
                   
                    data_appen = {**data_appen,"total_group": len(customers),"stt": idx_data}
                    
                    del data_appen["customers"]
                    new_data.append(data_appen)
                else:
                    new_data.append(customer)
        self.data_content = new_data
    # tạo nội dung bảng
    def make_table(self):
        
        sort_list = columnReport[self.report_type]["show_data"]
        # print("sort_list",sort_list)
        # Add the data to the worksheet
        start_row = columnReport[self.report_type]["content_start_at"]
        # chỉnh sửa , gom cột tự động- nhiều cột

        for row_idx, row_data in enumerate(self.data_content, start=start_row):
            # xử lý gộp cột nếu nhiều hơn 1 con
            if row_data.get("total_group") and row_data.get("total_group") > 1:
                total_group = row_data.get("total_group")
                column_group = list_name_col_x[0]
                group_string = f"{column_group}{row_idx}:{column_group}{row_idx+total_group-1}"
                self.ws.merge_cells(group_string)
                self.ws[f"{column_group}{row_idx+total_group-1}"].border = self.border_style
            # ghi stt vào cột stt
            if row_data.get("stt"):
                fields_group = self.fields_group or []
                cell = self.ws.cell(row=row_idx, column=1, value=row_data["stt"])
                cell.alignment = self.center_alignment
                cell.border = self.border_style
            for col_idx, value in enumerate(sort_list, start=2):
                # xử lý gộp cột nếu nhiều hơn 1 con
                if value in fields_group and  row_data.get("total_group") and row_data.get("total_group") >1:
                    total_group = row_data.get("total_group")
                    column_group = list_name_col_x[col_idx-1]
                    group_string = f"{column_group}{row_idx}:{column_group}{row_idx+total_group-1}"
                    self.ws.merge_cells(group_string)
                    self.ws[f"{column_group}{row_idx+total_group-1}"].border = self.border_style
                # ghi nội dung vào cột
                if value in fields_group and row_data.get("total_group") is not None:
                    cell = self.ws.cell(row=row_idx, column=col_idx, value=row_data.get(value) or "")
                    cell.alignment = self.left_alignment if col_idx <= 4 else self.center_alignment
                    cell.border = self.border_style
                if value not in fields_group:
                    cell = self.ws.cell(row=row_idx, column=col_idx, value=row_data.get(value) or "")
                    cell.alignment = self.left_alignment if col_idx <= 4 else self.center_alignment
                    cell.border = self.border_style
        return self
    pass

class MakeExcelInventory(MakeExcelCheckin):
    def __init__(self,report_type,data,from_time ="",to_time="",area="Tất cả"):
        if bool(from_time):
            self.from_time =  self.convert_tp_to_string(from_time)
        if bool(to_time):
            self.to_time =  self.convert_tp_to_string(to_time)
        self.area = area
        super().__init__(report_type,data,from_time,to_time)
        self.fields_group = ["customer_code","customer_name","customer_type","customer_address"]
        # self.changer_data()
    # định nghĩa thời gian
    def line_time(self):
        if hasattr(self, 'from_time') and hasattr(self, 'to_time'):
            return ["", "", f"Khu vực: {self.area}", "", "", ""],["", f"{self.from_time} - {self.to_time}", "", "", "", ""]
        elif hasattr(self, 'from_time'):
            return ["", "", f"Khu vực: {self.area}", "", "", ""],["", f"Từ {self.from_time}", "", "", "", ""]
        elif hasattr(self, 'to_time'):
            return ["", "", f"Khu vực: {self.area}", "", "", ""],["", f"Tính đến {self.to_time}", "", "", "", ""]
        else:
            return ["", "", f"Khu vực: {self.area}", "", "", ""]

    # biến đổi data
    def changer_data(self):
        data  = self.data_content
        new_data = []
        for idx, inven in enumerate(data,1):
            # xử lý dữ liệu 
            items = inven.get("items")
            if items:
                for index_item, item in enumerate(items):
                    #xử lý dữ liệu 
                    if item.get("update_at"):
                        item["update_at"] = self.convert_tp_to_string(item.get("update_at"))
                    if item.get("exp_time"):
                        item["exp_time"] = self.convert_tp_to_string(item.get("exp_time"))
                    if item.get("item_price"):
                        item["item_price"] = round(float(item.get("item_price")),2)
                        pass
                    # xử lý làm phẳng mảng, index=0 thì có thông tin chung
                    if index_item == 0 :
                        data_appen = {**inven,**item}
                    
                        data_appen = {**data_appen,"total_group": len(items),"stt": idx}
                        
                        del data_appen["items"]
                        new_data.append(data_appen)
                    else:
                        new_data.append(item)
            else:
                new_data.append(inven)
        self.data_content = new_data


class MakeExcelSell(MakeExcelCheckin):
    def __init__(self,report_type,data,from_time ="",to_time=""):
        super().__init__(report_type,data,from_time,to_time)
        self.fields_group = ['name', 'customer', 'territory', 'posting_date', 'sales_person','total', 'tax_amount', 'discount_amount', 'grand_total']
        # self.changer_data()
    def changer_data(self):
        data  = self.data_content
        # thêm "sum_qty","","","","sum_discount_items","sum_grand_items"
        data_footer = self.data_footer
        data_footer = {**data_footer,"sum_qty":0,"sum_discount_items":0,"sum_grand_items":0}
        new_data = []
        qty=0
        discount_amount=0
        amount=0
        for idx, sale in enumerate(data,1):
            # if "/" not in sale["posting_date"]:
            sale["posting_date"] = self.convert_tp_to_string(sale.get("posting_date"))
            # xử lý dữ liệu 
            items = sale.get("items")
            if items:
                for index_item, item in enumerate(items):
                    #xử lý dữ liệu 
                    qty += item["qty"]
                    discount_amount += item["discount_amount"]
                    amount += item["amount"]
                    # xử lý làm phẳng mảng, index=0 thì có thông tin chung
                    if index_item == 0 :
                        data_appen = {**sale,**item}
                    
                        data_appen = {**data_appen,"total_group": len(items),"stt": idx}
                        
                        del data_appen["items"]
                        new_data.append(data_appen)
                    else:
                        new_data.append(item)
            else:
                new_data.append(sale)
        data_footer["sum_qty"] = qty
        data_footer["sum_discount_items"] = discount_amount
        data_footer["sum_grand_items"] = amount
        self.data_content = new_data
        self.data_footer = data_footer

class MakeExcelOrder(MakeExcelSell):
    def __init__(self,report_type,data,from_time ="",to_time=""):
        super().__init__(report_type,data,from_time,to_time)
        self.fields_group = ['name', 'customer', 'territory', 'transaction_date', 'sales_person','total', 'tax_amount', 'discount_amount', 'grand_total']
    def changer_data(self):
        data  = self.data_content
        # thêm "sum_qty","","","","sum_discount_items","sum_grand_items"
        data_footer = self.data_footer
        data_footer = {**data_footer,"sum_qty":0,"sum_discount_items":0,"sum_grand_items":0}
        new_data = []
        qty=0
        discount_amount=0
        amount=0
        for idx, sale in enumerate(data,1):
            # if "/" not in sale["posting_date"]:
            sale["transaction_date"] = self.convert_tp_to_string(sale.get("transaction_date"))
            # xử lý dữ liệu 
            items = sale.get("items")
            if items:
                for index_item, item in enumerate(items):
                    #xử lý dữ liệu 
                    qty += item["qty"]
                    discount_amount += item["discount_amount"]
                    amount += item["amount"]
                    # xử lý làm phẳng mảng, index=0 thì có thông tin chung
                    if index_item == 0 :
                        data_appen = {**sale,**item}
                    
                        data_appen = {**data_appen,"total_group": len(items),"stt": idx}
                        
                        del data_appen["items"]
                        new_data.append(data_appen)
                    else:
                        new_data.append(item)
            else:
                new_data.append(sale)
        data_footer["sum_qty"] = qty
        data_footer["sum_discount_items"] = discount_amount
        data_footer["sum_grand_items"] = amount
        self.data_content = new_data
        self.data_footer = data_footer

class MakeExcelCustomer(MakeExcel):
    pass

class MakeExcelCustomerCheckin(MakeExcel):
    pass

